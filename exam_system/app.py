#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
病理学机考系统 - Flask 后端
参考 SurveyKing 文本导入功能设计
功能：注册/登录 | 试卷管理(增删改+发布/下线) | Excel/文本试题导入 | 随机抽题组卷 | 考试答题 | 成绩记录
"""

import os, hashlib, json, secrets, datetime, re, io, uuid
from functools import wraps
from flask import Flask, request, jsonify, send_from_directory
from flask_cors import CORS
import pymysql
from pymysql.cursors import DictCursor

app = Flask(__name__, static_folder='static', static_url_path='')
CORS(app, supports_credentials=True)
app.config['SECRET_KEY'] = secrets.token_hex(32)
app.config['MAX_CONTENT_LENGTH'] = 32 * 1024 * 1024  # 32MB

DB = {
    "host": "127.0.0.1", "port": 3306,
    "user": "root", "password": "Zfz19801210*",
    "database": "yf_boot_exam", "charset": "utf8mb4",
    "cursorclass": DictCursor,
}
REPO_ID = '3000000000000000001'

def db(): return pymysql.connect(**DB)

# ============ Auth ============
def hash_pwd(p, s):
    h = hashlib.md5(p.encode()).hexdigest()
    return hashlib.md5((h + s).encode()).hexdigest()

_token_key = app.config['SECRET_KEY']

def make_token(uid):
    import base64, time
    p = {"uid": uid, "exp": int(time.time()) + 86400 * 7}
    b = base64.urlsafe_b64encode(json.dumps(p).encode()).decode()
    return b + "." + hashlib.sha256((b + _token_key).encode()).hexdigest()

def check_token(tok):
    import base64, time
    try:
        a, s = tok.split('.')
        if s != hashlib.sha256((a + _token_key).encode()).hexdigest(): return
        p = json.loads(base64.urlsafe_b64decode(a + '==').decode())
        if p['exp'] < time.time(): return
        return p
    except: return

def auth(f):
    @wraps(f)
    def w(*a, **k):
        t = request.headers.get('Authorization', '')
        if t.startswith('Bearer '): t = t[7:]
        p = check_token(t)
        if not p: return jsonify({"code":401,"message":"请先登录"}),401
        request.uid = p['uid']
        return f(*a, **k)
    return w

def admin(f):
    @wraps(f)
    @auth
    def w(*a, **k):
        d = db()
        try:
            with d.cursor() as c:
                c.execute("SELECT user_name FROM el_sys_user WHERE id=%s",(request.uid,))
                u = c.fetchone()
                if not u or u['user_name'] != 'admin':
                    return jsonify({"code":403,"message":"需要管理员权限"}),403
        finally: d.close()
        return f(*a, **k)
    return w

# ============ Util ============
def clean(t):
    if not t: return ""
    t = re.sub(r'<[^>]+>','',t)
    return t.replace('&nbsp;',' ').replace('&lt;','<').replace('&gt;','>').replace('&amp;','&').strip()

def new_id():
    return str(int(datetime.datetime.now().timestamp() * 1000000))

# ============ Auth API ============
@app.route('/api/register', methods=['POST'])
def register():
    j = request.get_json()
    u = (j.get('username') or '').strip()
    p = (j.get('password') or '').strip()
    if not u or len(p) < 4:
        return jsonify({"code":400,"message":"用户名不能为空，密码至少4位"}),400
    d = db()
    try:
        with d.cursor() as c:
            c.execute("SELECT id FROM el_sys_user WHERE user_name=%s",(u,))
            if c.fetchone(): return jsonify({"code":400,"message":"用户名已存在"}),400
            uid = new_id()
            s = secrets.token_hex(3)
            hp = hash_pwd(p, s)
            c.execute("""INSERT INTO el_sys_user (id,user_name,real_name,password,salt,state,create_time,update_time)
                VALUES (%s,%s,%s,%s,%s,1,NOW(),NOW())""",(uid,u,u,hp,s))
            d.commit()
        return jsonify({"code":200,"message":"注册成功"})
    except Exception as e: return jsonify({"code":500,"message":str(e)}),500
    finally: d.close()

@app.route('/api/login', methods=['POST'])
def login():
    j = request.get_json()
    u = (j.get('username') or '').strip()
    p = (j.get('password') or '').strip()
    if not u or not p: return jsonify({"code":400,"message":"请输入用户名和密码"}),400
    d = db()
    try:
        with d.cursor() as c:
            c.execute("SELECT id,user_name,real_name,password,salt FROM el_sys_user WHERE user_name=%s",(u,))
            r = c.fetchone()
            if not r or hash_pwd(p, r['salt']) != r['password']:
                return jsonify({"code":401,"message":"用户名或密码错误"}),401
            return jsonify({"code":200,"data":{
                "token": make_token(r['id']), "userId": r['id'],
                "username": r['user_name'],
                "realName": r['real_name'] or r['user_name'],
                "isAdmin": r['user_name'] == 'admin'
            }})
    except Exception as e: return jsonify({"code":500,"message":str(e)}),500
    finally: d.close()

# ============ Paper API ============
@app.route('/api/papers', methods=['GET'])
@auth
def paper_list():
    d = db()
    try:
        with d.cursor() as c:
            c.execute("SELECT * FROM exam_paper WHERE status=1 ORDER BY create_time DESC")
            ps = c.fetchall()
        for p in ps:
            p['totalQuestions'] = p['radio_count']+p['multi_count']+p['material_count']+p['case_count']
        return jsonify({"code":200,"data":ps})
    finally: d.close()

@app.route('/api/admin/papers', methods=['GET'])
@admin
def admin_papers():
    d = db()
    try:
        with d.cursor() as c:
            c.execute("SELECT * FROM exam_paper ORDER BY create_time DESC")
            ps = c.fetchall()
        for p in ps:
            p['totalQuestions'] = p['radio_count']+p['multi_count']+p['material_count']+p['case_count']
        return jsonify({"code":200,"data":ps})
    finally: d.close()

@app.route('/api/admin/papers', methods=['POST'])
@admin
def paper_create():
    j = request.get_json()
    pid = j.get('id') or ('p'+new_id())
    d = db()
    try:
        with d.cursor() as c:
            c.execute("""INSERT INTO exam_paper (id,title,duration,radio_count,multi_count,material_count,case_count,total_score,status)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
                (pid, j.get('title','新试卷'), j.get('duration',120),
                 j.get('radio_count',30), j.get('multi_count',20),
                 j.get('material_count',20), j.get('case_count',30),
                 j.get('total_score',100), j.get('status',1)))
            d.commit()
        return jsonify({"code":200,"message":"创建成功","data":{"id":pid}})
    except Exception as e: return jsonify({"code":500,"message":str(e)}),500
    finally: d.close()

@app.route('/api/admin/papers/<pid>', methods=['PUT'])
@admin
def paper_update(pid):
    j = request.get_json()
    d = db()
    try:
        fs, vs = [], []
        for k in ['title','duration','radio_count','multi_count','material_count','case_count','total_score','status']:
            if k in j: fs.append(f"{k}=%s"); vs.append(j[k])
        if fs:
            vs.append(pid)
            with d.cursor() as c:
                c.execute(f"UPDATE exam_paper SET {','.join(fs)},update_time=NOW() WHERE id=%s", vs)
                d.commit()
        return jsonify({"code":200,"message":"更新成功"})
    except Exception as e: return jsonify({"code":500,"message":str(e)}),500
    finally: d.close()

@app.route('/api/admin/papers/<pid>', methods=['DELETE'])
@admin
def paper_delete(pid):
    d = db()
    try:
        with d.cursor() as c:
            c.execute("DELETE FROM exam_paper WHERE id=%s",(pid,))
            d.commit()
        return jsonify({"code":200,"message":"删除成功"})
    except Exception as e: return jsonify({"code":500,"message":str(e)}),500
    finally: d.close()

# ============ Question Stats ============
@app.route('/api/admin/questions/counts', methods=['GET'])
@admin
def question_counts():
    d = db()
    try:
        with d.cursor() as c:
            c.execute("SELECT qu_type, COUNT(*) as cnt FROM el_repo_qu WHERE repo_id=%s GROUP BY qu_type",(REPO_ID,))
            rs = {r['qu_type']:r['cnt'] for r in c.fetchall()}
        return jsonify({"code":200,"data":rs})
    finally: d.close()

@app.route('/api/admin/questions', methods=['GET'])
@admin
def question_list():
    qt = request.args.get('type','')
    pg = int(request.args.get('page',1))
    lm = int(request.args.get('limit',30))
    off = (pg-1)*lm
    d = db()
    try:
        with d.cursor() as c:
            wh = "WHERE repo_id=%s" + (" AND qu_type=%s" if qt else "")
            args = [REPO_ID] + ([qt] if qt else [])
            c.execute(f"SELECT COUNT(*) as total FROM el_repo_qu {wh}", args)
            total = c.fetchone()['total']
            c.execute(f"SELECT id,content,qu_type,analysis FROM el_repo_qu {wh} ORDER BY qu_type,id LIMIT %s OFFSET %s",
                       args+[lm,off])
            qs = c.fetchall()
        return jsonify({"code":200,"data":{"total":total,"page":pg,"items":qs}})
    finally: d.close()

# ============ 试题导入（参考 SurveyKing） ============

# SurveyKing Excel 列名映射
SURVEYKING_HEADERS = {
    '序号': 'serial', '题干': 'content', '题目': 'content',
    '选项A': 'optA', '选项B': 'optB', '选项C': 'optC', '选项D': 'optD',
    '选项E': 'optE', '选项F': 'optF', '选项G': 'optG', '选项H': 'optH',
    '解析': 'analysis', '答案解析': 'analysis', '答案': 'answer',
    '分数': 'score', '分值': 'score', '标签': 'tags', 'tag': 'tags',
}

SHEET_TYPE_MAP = {
    '单选题': 'radio', '单选': 'radio',
    '多选题': 'multi', '多选': 'multi',
    '共用题干题': 'material', '共用题干': 'material', '材料题': 'material',
    '案例分析题': 'case', '案例分析': 'case',
}

def _parse_import_row(row, qu_type):
    """解析一行导入数据，返回题目+选项列表"""
    content = row.get('content', '')
    if not content: return None

    # 收集选项
    options = []
    answer_raw = row.get('answer', '').strip().upper()
    correct_set = set()
    if answer_raw:
        # 支持 A,B,C 或 ABC 格式
        parts = re.findall(r'[A-H]', answer_raw)
        correct_set = set(parts)

    for letter in 'ABCDEFGH':
        opt_text = row.get(f'opt{letter}', '')
        if opt_text and opt_text.strip():
            is_right = 1 if letter in correct_set else 0
            options.append({'tag': letter, 'content': opt_text.strip(), 'is_right': is_right})

    if not options:
        # 没有选项，尝试生成默认选项
        default_opts = [
            {'tag': 'A', 'content': '正确', 'is_right': 1 if 'A' in correct_set else 0},
            {'tag': 'B', 'content': '错误', 'is_right': 1 if 'B' in correct_set else 0},
        ]
        options = default_opts if qu_type in ('radio','multi','material','case') else []

    analysis = clean(row.get('analysis', ''))

    return {
        'content': clean(content),
        'analysis': analysis,
        'options': options,
        'qu_type': qu_type,
    }

@app.route('/api/admin/import/excel', methods=['POST'])
@admin
def import_excel():
    """参考 SurveyKing 格式导入 Excel 试题"""
    if 'file' not in request.files:
        return jsonify({"code":400,"message":"请上传文件"}),400

    file = request.files['file']
    if not file.filename.endswith(('.xlsx','.xls')):
        return jsonify({"code":400,"message":"请上传 .xlsx 或 .xls 格式文件"}),400

    try:
        import openpyxl
        wb = openpyxl.load_workbook(file, read_only=True)
    except ImportError:
        return jsonify({"code":500,"message":"服务器未安装openpyxl，请联系管理员"}),500
    except Exception as e:
        return jsonify({"code":400,"message":f"文件读取失败: {str(e)}"}),400

    total_imported = 0
    total_skipped = 0
    errors = []

    for sheet_name in wb.sheetnames:
        qu_type = SHEET_TYPE_MAP.get(sheet_name)
        if not qu_type:
            continue  # 跳过不认识的 sheet

        ws = wb[sheet_name]
        # 读取表头
        headers = []
        first_row = True
        for row in ws.iter_rows(values_only=True):
            if first_row:
                headers = [str(h).strip() if h else '' for h in row]
                first_row = False
                continue

            if not row or not any(row):  # 空行跳过
                continue

            # 构建行数据字典
            row_data = {}
            for i, val in enumerate(row):
                if i < len(headers) and headers[i]:
                    col_name = SURVEYKING_HEADERS.get(headers[i], headers[i])
                    row_data[col_name] = str(val).strip() if val is not None else ''

            parsed = _parse_import_row(row_data, qu_type)
            if not parsed or not parsed['content']:
                total_skipped += 1
                continue

            # 插入数据库
            try:
                _insert_question(parsed)
                total_imported += 1
            except Exception as e:
                errors.append(f"Sheet[{sheet_name}] 行导入失败: {str(e)}")
                total_skipped += 1

    wb.close()
    return jsonify({
        "code": 200,
        "message": f"导入完成：成功 {total_imported} 题，跳过 {total_skipped} 题",
        "data": {"imported": total_imported, "skipped": total_skipped, "errors": errors[:10]}
    })

@app.route('/api/admin/import/text', methods=['POST'])
@admin
def import_text():
    """导入文本试题 - 支持完整层级格式
    格式：
    【单选题】
    分类：各论-消化
    8. 题目内容
    A. 选项A
    ...
    答案：G
    解析：解析内容

    【共用题干】/ 【案例分析题】
    题干：共享题干内容
    提问1：子题内容
    A. ...
    答案：C
    解析：...
    提问2：...
    """
    j = request.get_json()
    default_type = j.get('type', 'radio')
    text = j.get('text', '').strip()

    if not text:
        return jsonify({"code":400,"message":"请输入试题内容"}),400

    imported, skipped = 0, 0

    # 按【题型】标记分段
    sections = re.split(r'【([^】]+)】', text)

    # sections[0] is text before first 【】, then alternating: type_name, content, type_name, content...
    i = 1  # skip leading text before first type marker
    while i < len(sections) - 1:
        type_name = sections[i].strip()
        section_text = sections[i + 1].strip()
        i += 2

        qu_type = SHEET_TYPE_MAP.get(type_name, default_type)
        is_grouped = qu_type in ('material', 'case')

        # 按 分类：开头的块或连续空行分割大题组
        groups = _split_question_groups(section_text, is_grouped)

        for group in groups:
            try:
                count = _import_question_group(group, qu_type)
                imported += count
            except Exception as e:
                import traceback
                print(f"[Import Error] {str(e)}", flush=True)
                traceback.print_exc()
                skipped += 1

    return jsonify({
        "code": 200,
        "message": f"导入完成：成功 {imported} 题，跳过 {skipped} 组",
        "data": {"imported": imported, "skipped": skipped}
    })


def _split_question_groups(text, is_grouped):
    """将文本段拆分为独立的题目组。
    对于共用题干/案例分析：每个 题干： 开头为一组。
    对于单选/多选：每个带编号的题目为一组。
    """
    lines = text.split('\n')
    groups = []
    current = []

    for line in lines:
        stripped = line.strip()
        # 检测新组的开始标记
        is_new_group = False
        if is_grouped and re.match(r'^题干[：:]', stripped):
            is_new_group = True
        elif not is_grouped and re.match(r'^(\d+)[.、)\s]', stripped) and current:
            # 单选/多选：遇到新的编号题目，且当前已有内容
            is_new_group = True

        if is_new_group and current:
            groups.append('\n'.join(current))
            current = [line]
        else:
            if stripped or current:  # 保持内容连贯
                current.append(line)

    if current:
        joined = '\n'.join(current).strip()
        if joined:
            groups.append(joined)

    return groups


def _import_question_group(text, qu_type):
    """导入一组题目，返回导入的题目数。
    对于共用题干/案例分析：解析 题干 + 提问N 结构。
    对于单选/多选：解析单道题。
    """
    is_grouped = qu_type in ('material', 'case')

    if not is_grouped:
        # 单选/多选：直接解析
        q = _parse_single_question(text, qu_type)
        if q:
            _insert_question(q)
            return 1
        return 0

    # 共用题干/案例分析：提取共享题干和各子题
    lines = [l for l in text.split('\n')]
    stem = ''
    category = ''
    sub_questions = []  # [(sub_title, sub_lines)]
    current_sub = None
    current_lines = []

    for line in lines:
        stripped = line.strip()

        # 提取分类
        m_cat = re.match(r'^分类[：:]\s*(.+)', stripped)
        if m_cat:
            category = m_cat.group(1).strip()
            continue

        # 提取题干
        m_stem = re.match(r'^题干[：:]\s*(.*)', stripped)
        if m_stem:
            stem = m_stem.group(1).strip()
            # 题干可能跨多行，后续非 提问 的行也属于题干
            continue

        # 提取提问标记
        m_sub = re.match(r'^提问(\d+)[：:]\s*(.*)', stripped)
        if m_sub:
            # 保存之前的子题
            if current_sub is not None:
                sub_questions.append((current_sub, current_lines))
            current_sub = m_sub.group(2).strip() or f'提问{m_sub.group(1)}'
            current_lines = []
            continue

        # 如果已经有子题在解析中，累积到当前子题
        if current_sub is not None:
            current_lines.append(line)
        elif stem and stripped:
            # 题干后续内容（多行题干）
            stem += '\n' + stripped

    # 保存最后一个子题
    if current_sub is not None:
        sub_questions.append((current_sub, current_lines))

    if not sub_questions:
        # 没有子题结构，当作单题处理
        q = _parse_single_question(text, qu_type)
        if q:
            _insert_question(q)
            return 1
        return 0

    # 为每个子题生成独立记录，content = 题干 + 子题
    count = 0
    for sub_title, sub_lines in sub_questions:
        sub_text = '\n'.join(sub_lines)
        full_content = f"{stem}\n\n{sub_title}" if stem else sub_title
        q = _parse_single_question(full_content + '\n' + sub_text, qu_type)
        if q:
            # 用完整题干替换 content
            q['content'] = clean(full_content)
            _insert_question(q)
            count += 1

    return count


def _parse_single_question(text, qu_type):
    """解析单道题目文本，返回 content, options, analysis, qu_type"""
    lines = [l.strip() for l in text.split('\n') if l.strip()]
    if not lines:
        return None

    content = ''
    options = []
    answer_set = set()
    analysis = ''

    i = 0

    # 1. 解析题目内容（选项行之前的所有文本）
    while i < len(lines):
        ln = lines[i]
        if re.match(r'^[A-H][.、)\s]', ln):
            break
        if ln.startswith('答案') or ln.startswith('解析'):
            break
        content += (ln + '\n') if content else ln
        i += 1
    content = content.strip()

    # 2. 解析选项
    while i < len(lines):
        ln = lines[i]
        m = re.match(r'^([A-H])[.、)\s]+(.+)', ln, re.IGNORECASE)
        if m:
            options.append({'tag': m.group(1).upper(), 'content': m.group(2).strip(), 'is_right': 0})
            i += 1
        else:
            break

    # 3. 解析答案
    if i < len(lines) and lines[i].startswith('答案'):
        ans_raw = re.sub(r'^答案[：:]*\s*', '', lines[i]).strip()
        parts = re.findall(r'[A-H]', ans_raw.upper())
        answer_set = set(parts)
        i += 1

    # 4. 解析答案（多行情况：答案跨行）
    while i < len(lines) and re.match(r'^[A-H]+$', lines[i].strip()):
        parts = re.findall(r'[A-H]', lines[i].strip().upper())
        answer_set.update(parts)
        i += 1

    # 5. 解析解析（剩余所有行）
    while i < len(lines):
        ln = lines[i]
        # 去掉 "解析：" 或 "解析:" 前缀
        ln = re.sub(r'^解析[：:]\s*', '', ln)
        analysis += (ln + '\n') if analysis else ln
        i += 1
    analysis = analysis.strip()

    # 标记正确选项
    for opt in options:
        if opt['tag'] in answer_set:
            opt['is_right'] = 1

    # 无选项时生成默认
    if not options:
        for letter in 'ABCD':
            options.append({'tag': letter, 'content': f'选项{letter}', 'is_right': 1 if letter in answer_set else 0})

    if not content:
        return None

    return {
        'content': content,
        'analysis': analysis,
        'options': options,
        'qu_type': qu_type,
    }

def _insert_question(q):
    """插入一道试题到数据库"""
    d = db()
    try:
        qid = new_id()
        with d.cursor() as c:
            c.execute("""INSERT INTO el_repo_qu (id,content,analysis,qu_type,repo_id)
                VALUES (%s,%s,%s,%s,%s)""",
                (qid, q['content'], q.get('analysis',''), q['qu_type'], REPO_ID))
            for opt in q.get('options', []):
                c.execute("""INSERT INTO el_repo_qu_answer (id,qu_id,tag,content,is_right)
                    VALUES (%s,%s,%s,%s,%s)""",
                    (new_id(), qid, opt['tag'], opt['content'], opt.get('is_right', 0)))
            d.commit()
    finally: d.close()

# ============ 考试 API ============
@app.route('/api/exam/<paper_id>', methods=['GET'])
@auth
def start_exam(paper_id):
    """获取试卷内容（随机抽题）"""
    d = db()
    try:
        # 获取试卷配置
        with d.cursor() as c:
            c.execute("SELECT * FROM exam_paper WHERE id=%s",(paper_id,))
            paper = c.fetchone()
        if not paper:
            return jsonify({"code":404,"message":"试卷不存在"}),404

        sections = []

        configs = [
            ('radio', '单选题', paper['radio_count']),
            ('multi', '多选题', paper['multi_count']),
            ('material', '共用题干题', paper['material_count']),
            ('case', '案例分析题', paper['case_count']),
        ]

        for qu_type, name, count in configs:
            if count <= 0: continue
            qs = _pick_questions(d, qu_type, count, user_id=request.uid)
            if qs:
                sections.append({
                    "type": qu_type, "name": name,
                    "desc": f"共{len(qs)}题",
                    "questions": qs,
                })

        return jsonify({
            "code": 200,
            "data": {
                "id": paper_id,
                "title": paper['title'],
                "duration": paper['duration'],
                "total_score": paper['total_score'],
                "sections": sections,
            }
        })
    except Exception as e:
        return jsonify({"code":500,"message":str(e)}),500
    finally: d.close()

def _pick_questions(d, qu_type, limit, user_id=None, paper_id=None):
    """随机抽取指定类型和数量的题目，避免同一用户上一张卷子中的题"""
    with d.cursor() as c:
        # 查用户最近一次考试做过的题目ID（只限同repo同类型）
        seen_ids = set()
        if user_id:
            c.execute("""SELECT answers FROM exam_record WHERE user_id=%s
                ORDER BY create_time DESC LIMIT 1""", (user_id,))
            row = c.fetchone()
            if row:
                try:
                    ans = json.loads(row['answers'])
                    seen_ids = set(ans.keys())
                except: pass

        # 查该类型的总可用题目
        c.execute("SELECT COUNT(*) as cnt FROM el_repo_qu WHERE qu_type=%s AND repo_id=%s", (qu_type, REPO_ID))
        total = c.fetchone()['cnt']
        needed = min(limit, max(total, limit))  # 不超总数

        if seen_ids and total > len(seen_ids):
            # 尽量排除最近做过的
            placeholders = ','.join(['%s']*len(seen_ids))
            try:
                c.execute(f"""SELECT id,content,analysis,qu_type FROM el_repo_qu
                    WHERE qu_type=%s AND repo_id=%s AND id NOT IN ({placeholders})
                    ORDER BY RAND() LIMIT %s""",
                    (qu_type, REPO_ID, *seen_ids, needed))
                qs = c.fetchall()
                if len(qs) >= needed * 0.6:  # 排除后还够60%就用
                    pass
                else:
                    raise Exception("not enough")
            except:
                c.execute("""SELECT id,content,analysis,qu_type FROM el_repo_qu
                    WHERE qu_type=%s AND repo_id=%s ORDER BY RAND() LIMIT %s""",
                    (qu_type, REPO_ID, needed))
                qs = c.fetchall()
        else:
            c.execute("""SELECT id,content,analysis,qu_type FROM el_repo_qu
                WHERE qu_type=%s AND repo_id=%s ORDER BY RAND() LIMIT %s""",
                (qu_type, REPO_ID, needed))
            qs = c.fetchall()

    result = []
    for q in qs:
        with d.cursor() as c:
            c.execute("SELECT tag,content,is_right FROM el_repo_qu_answer WHERE qu_id=%s ORDER BY tag",(q['id'],))
            answers = c.fetchall()
        options = []
        correct = []
        for a in answers:
            options.append({"key":a['tag'],"text":clean(a['content'])})
            if a['is_right'] == 1: correct.append(a['tag'])
        result.append({
            "id": q['id'], "content": clean(q['content']),
            "options": options, "correct": correct,
            "analysis": clean(q.get('analysis','')), "type": q['qu_type'],
        })
    return result

@app.route('/api/exam/<paper_id>/submit', methods=['POST'])
@auth
def submit_exam(paper_id):
    """提交试卷答案，返回评分+全部题目详情用于回顾
    计分规则：
    - 单选题(radio)：每题1分，错选不得分
    - 共用题干题(material)：每小问1分，错选不得分
    - 多选题(multi)：每题2分，必须与标准答案完全一致才得分
    - 案例分析题(case)：每个提问1分，按选项权重得分/扣分，最低0分
    """
    j = request.get_json()
    answers = j.get('answers', {})
    d = db()
    try:
        correct = 0; total = 0; score = 0.0; max_score = 0.0; details = []
        for qid, ua in answers.items():
            total += 1
            if isinstance(ua, str): ua = [ua]
            ua_sorted = sorted(ua)
            with d.cursor() as c:
                c.execute("SELECT tag,content,is_right FROM el_repo_qu_answer WHERE qu_id=%s ORDER BY tag",(qid,))
                opts = c.fetchall()
            ct = sorted([r['tag'] for r in opts if r['is_right'] == 1])

            # 查题目信息
            with d.cursor() as c:
                c.execute("SELECT content,analysis,qu_type FROM el_repo_qu WHERE id=%s",(qid,))
                qinfo = c.fetchone()

            qu_type = qinfo['qu_type'] if qinfo else ''
            is_c = ua_sorted == ct
            if is_c: correct += 1

            # ---- 分题型计分 ----
            q_score = 0.0
            q_max = 1.0

            if qu_type == 'radio':
                # 单选题：每题1分，错选不得分
                q_max = 1.0
                q_score = 1.0 if is_c else 0.0

            elif qu_type == 'material':
                # 共用题干单选题：每小问1分，错选不得分
                q_max = 1.0
                q_score = 1.0 if is_c else 0.0

            elif qu_type == 'multi':
                # 多选题：每题2分，必须与标准答案完全一致才得分
                q_max = 2.0
                q_score = 2.0 if is_c else 0.0

            elif qu_type == 'case':
                # 案例分析题：每个提问1分，按选项权重得分/扣分
                q_max = 1.0
                n_correct = len(ct)
                if n_correct > 0:
                    w = 1.0 / n_correct  # 每个正确选项的分值权重
                    # 少选：按选对的选项得对应分值
                    correct_picks = set(ua_sorted) & set(ct)
                    q_score = len(correct_picks) * w
                    # 多选/错选：倒扣对应选项的分值
                    wrong_picks = set(ua_sorted) - set(ct)
                    q_score -= len(wrong_picks) * w
                    # 保底：最低0分
                    q_score = max(0.0, q_score)
                else:
                    q_score = 0.0

            score += q_score
            max_score += q_max

            details.append({
                "quId": qid,
                "content": clean(qinfo['content']) if qinfo else '',
                "type": qu_type,
                "analysis": clean(qinfo['analysis']) if qinfo else '',
                "options": [{"key":r['tag'],"text":clean(r['content']),"isRight":r['is_right']==1} for r in opts],
                "userAnswer": ua_sorted,
                "correctAnswer": ct,
                "isCorrect": is_c,
                "questionScore": round(q_score, 2),
                "questionMax": q_max,
            })

        score = round(score, 2)
        max_score = round(max_score, 2)
        pct = round(score / max_score * 100, 1) if max_score > 0 else 0

        rid = new_id()
        try:
            with d.cursor() as c:
                c.execute("""INSERT INTO exam_record (id,user_id,paper_id,answers,score,total,correct_count,percentage,duration,create_time)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,NOW())""",
                    (rid, request.uid, paper_id, json.dumps(answers,ensure_ascii=False),
                     score, total, correct, pct, j.get('duration',0)))
            d.commit()
        except Exception as e:
            app.logger.warning(f"Failed to save exam record: {e}")

        return jsonify({"code":200,"data":{
            "totalCount":total,"correctCount":correct,
            "score":score,"maxScore":max_score,"percentage":pct,
            "details":details
        }})
    except Exception as e:
        return jsonify({"code":500,"message":str(e)}),500
    finally: d.close()

# ============ 成绩记录 ============
@app.route('/api/records', methods=['GET'])
@auth
def my_records():
    d = db()
    try:
        with d.cursor() as c:
            c.execute("""SELECT r.*, p.title as paper_title FROM exam_record r
                LEFT JOIN exam_paper p ON r.paper_id COLLATE utf8mb4_general_ci = p.id COLLATE utf8mb4_general_ci
                WHERE r.user_id=%s ORDER BY r.create_time DESC LIMIT 50""",(request.uid,))
            rs = c.fetchall()
        return jsonify({"code":200,"data":rs})
    finally: d.close()

@app.route('/api/admin/records', methods=['GET'])
@admin
def all_records():
    d = db()
    try:
        with d.cursor() as c:
            c.execute("""SELECT r.*, p.title as paper_title, u.user_name, u.real_name FROM exam_record r
                LEFT JOIN exam_paper p ON r.paper_id COLLATE utf8mb4_general_ci = p.id COLLATE utf8mb4_general_ci
                LEFT JOIN el_sys_user u ON r.user_id COLLATE utf8mb4_general_ci = u.id COLLATE utf8mb4_general_ci
                ORDER BY r.create_time DESC LIMIT 200""")
            rs = c.fetchall()
        return jsonify({"code":200,"data":rs})
    finally: d.close()

# ============ 用户管理 ============
@app.route('/api/admin/users', methods=['GET'])
@admin
def admin_users():
    """获取所有用户列表"""
    d = db()
    try:
        with d.cursor() as c:
            c.execute("""SELECT u.id, u.user_name, u.real_name, u.state, u.create_time,
                COUNT(r.id) as exam_count
                FROM el_sys_user u
                LEFT JOIN exam_record r ON r.user_id = u.id
                GROUP BY u.id, u.user_name, u.real_name, u.state, u.create_time
                ORDER BY u.create_time DESC""")
            us = c.fetchall()
        return jsonify({"code":200,"data":us})
    finally: d.close()

@app.route('/api/admin/users', methods=['DELETE'])
@admin
def admin_delete_users():
    """批量删除用户（不能删除admin）"""
    j = request.get_json()
    ids = j.get('ids', [])
    if not ids:
        return jsonify({"code":400,"message":"请选择要删除的用户"}),400

    d = db()
    try:
        # 过滤掉 admin 账号
        placeholders = ','.join(['%s'] * len(ids))
        with d.cursor() as c:
            c.execute(f"SELECT id, user_name FROM el_sys_user WHERE id IN ({placeholders})", ids)
            users = c.fetchall()

        to_delete = [u['id'] for u in users if u['user_name'] != 'admin']
        skipped = len(ids) - len(to_delete)

        if not to_delete:
            return jsonify({"code":400,"message":"不能删除管理员账号"}),400

        # 删除用户的考试记录
        with d.cursor() as c:
            ph = ','.join(['%s'] * len(to_delete))
            c.execute(f"DELETE FROM exam_record WHERE user_id IN ({ph})", to_delete)
            c.execute(f"DELETE FROM el_sys_user WHERE id IN ({ph})", to_delete)
            d.commit()

        msg = f"成功删除 {len(to_delete)} 个用户"
        if skipped:
            msg += f"，跳过 {skipped} 个（管理员账号）"
        return jsonify({"code":200,"message":msg})
    except Exception as e:
        return jsonify({"code":500,"message":str(e)}),500
    finally: d.close()

# ============ 题库模板下载（参考SurveyKing） ============
@app.route('/api/admin/template', methods=['GET'])
@admin
def download_template():
    """下载 SurveyKing 格式的 Excel 导入模板"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return jsonify({"code":500,"message":"服务器未安装openpyxl"}),500

    wb = openpyxl.Workbook()
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')

    sheets_config = [
        ('单选题', ['序号', '题干', '选项A', '选项B', '选项C', '选项D', '选项E', '选项F', '选项G', '选项H', '解析', '分数', '答案', '标签']),
        ('多选题', ['序号', '题干', '选项A', '选项B', '选项C', '选项D', '选项E', '选项F', '选项G', '选项H', '解析', '分数', '答案', '标签']),
        ('共用题干题', ['序号', '题干', '选项A', '选项B', '选项C', '选项D', '选项E', '选项F', '选项G', '选项H', '解析', '分数', '答案', '标签']),
        ('案例分析题', ['序号', '题干', '选项A', '选项B', '选项C', '选项D', '选项E', '选项F', '选项G', '选项H', '解析', '分数', '答案', '标签']),
    ]

    for sheet_name, headers in sheets_config:
        ws = wb.create_sheet(title=sheet_name)
        for i, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=i, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        # 设置列宽
        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 50
        for col in 'CDEFGHIJ':
            ws.column_dimensions[col].width = 15
        ws.column_dimensions['K'].width = 40
        ws.column_dimensions['L'].width = 8
        ws.column_dimensions['M'].width = 10
        ws.column_dimensions['N'].width = 15

        # 添加示例数据
        examples = {
            '单选题': ['1', '以下哪项是恶性肿瘤的特征？', '分化好', '异型性小', '浸润性生长', '核分裂象少', '', '', '', '', '恶性肿瘤的特征包括浸润性生长和转移。', '1', 'C', '肿瘤,病理'],
            '多选题': ['1', '以下哪些属于癌前病变？', '乳腺导管上皮增生', '慢性萎缩性胃炎', '结肠腺瘤', '宫颈上皮内瘤变', '', '', '', '', '癌前病变是指具有癌变潜能的良性病变。', '2', 'ABCD', '癌前病变'],
            '共用题干题': ['1', '患者男性，60岁。因上腹部不适就诊。', '胃镜活检', 'CT检查', 'B超', '血常规', '', '', '', '', '胃癌的临床表现和诊断方法。', '1', 'A', '临床案例'],
            '案例分析题': ['1', '患者女性，45岁。乳腺包块。以下哪些选项正确？', '考虑乳腺癌', '需做超声检查', '建议观察', '需做穿刺活检', '', '', '', '', '乳腺癌的诊断流程。', '1', 'ABD', '乳腺癌'],
        }
        ex = examples.get(sheet_name)
        if ex:
            for i, v in enumerate(ex, 1):
                ws.cell(row=2, column=i, value=v)

    # 删除默认sheet
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    wb.close()

    from flask import Response
    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': "attachment; filename*=UTF-8''%E8%AF%95%E9%A2%98%E5%AF%BC%E5%85%A5%E6%A8%A1%E6%9D%BF.xlsx"}
    )

# ============ Static Files ============
@app.route('/admin')
def admin_page():
    return send_from_directory(app.static_folder, 'admin.html')

@app.route('/')
def index():
    return send_from_directory(app.static_folder, 'index.html')

@app.route('/<path:path>')
def static_files(path):
    if path == 'admin': return send_from_directory(app.static_folder, 'admin.html')
    fp = os.path.join(app.static_folder, path)
    if os.path.exists(fp):
        return send_from_directory(app.static_folder, path)
    return send_from_directory(app.static_folder, 'index.html')

# ============ Startup ============
if __name__ == '__main__':
    os.makedirs(os.path.join(os.path.dirname(__file__), 'static'), exist_ok=True)
    print("=" * 50)
    print("病理学机考系统 v2 启动")
    print(f"端口: 8090")
    print("=" * 50)
    app.run(host='0.0.0.0', port=8090, debug=False)
